from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


wb = load_workbook('French Section A Internal Assessment Grades 2022.xlsx')
ws = wb.active

i = 1
for cell in range(3, 33):
    ws["B" + str(cell)] = "Student" + str(i)
    i += 1


print(wb.sheetnames)
ws = wb["French A_Internals Breakdown"]

i = 1
for cell in range(3, 33):
    ws["B" + str(cell)] = "Student" + str(i)
    i += 1

wb.save("French Section A Internal Assessment Grades 2022.xlsx")


print(ws)
print(ws['A1'].value)

print(wb.sheetnames)
wb.create_sheet("Test")
print(wb.sheetnames)
wb["Test"].title = "TestNew"
print(wb.sheetnames)


ws.title = "French Section A"
print(wb.sheetnames)

wb.save("French Section A Internal Assessment Grades 2022.xlsx")

wb = load_workbook("French Section A Internal Assessment Grades 2022.xlsx")
ws = wb.active

for row in range(1, 35):
    for col in range(1, 3):
        cell_name = get_column_letter(col)
        print(ws[cell_name + str(row)])
        print(ws[cell_name + str(row)].value)

ws.unmerge_cells("A1:D1")
ws.merge_cells("A1:D1")
ws.insert_rows(7)
ws.insert_cols(2)
ws.delete_rows(7)
ws.delete_cols(2)

ws.move_range("C1:D31", rows=-1, cols=5)  # minus moves up/left, plus moves down/right

wb.save("French Section A Internal Assessment Grades 2022.xlsx")


wb = Workbook()
ws = wb.active
ws.title = "Data"

ws.append(['Coding', 'Is', 'Awesome', 'Yo'])
ws.append(['Coding', 'Is', 'Awesome', 'And', 'You', 'Know', 'It'])
ws.append(['micdrop'])

wb.save("Coding.xlsx")

data = {
    "Joseph Priestly": {
        "Math": 88,
        "Science": 87,
        "Computers": 91,
        "English": 81,
        "Gym": 95,
    },
    "Pierre Curie": {
        "Math": 93,
        "Science": 89,
        "Computers": 84,
        "English": 92,
        "Gym": 87,
    },
    "Marie Curie": {
        "Math": 81,
        "Science": 83,
        "Computers": 88,
        "English": 99,
        "Gym": 80,
    },
    "Antoine Lavoisier": {
        "Math": 79,
        "Science": 80,
        "Computers": 80,
        "English": 98,
        "Gym": 81,
    },
    "Charles Darwin": {
        "Math": 98,
        "Science": 92,
        "Computers": 93,
        "English": 87,
        "Gym": 100,
    },
    "Isaac Newton": {
        "Math": 93,
        "Science": 79,
        "Computers": 90,
        "English": 91,
        "Gym": 88,
    },
    "Nikola Tesla": {
        "Math": 88,
        "Science": 95,
        "Computers": 80,
        "English": 90,
        "Gym": 79,
    }
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ["Names"] + list(data["Joseph Priestly"].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

wb.save("New Grades.xlsx")


wb = load_workbook("New Grades.xlsx")
ws = wb.active

print(ws.max_column)
print(ws.max_row)

for col in range(2, len(data['Joseph Priestly']) + 2):
    char = get_column_letter(col)
    ws[char + str(9)] = "=AVERAGE(" + char + str(2) + ":" + char + str(8) + ")"

wb.save("New Grades.xlsx")
